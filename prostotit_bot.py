import datetime
import re

import telebot
import openpyxl
import os
import schedule
import message_handler
import time

from dateutil.relativedelta import relativedelta

from tooken import token
from datetime import datetime, timedelta

from utils import wright_name, wright_time, wright_date, wright_month, wright_week, wright_days, ret_urn_day, \
    wright_last_days, wright_delete, wright_chat_id, wright_text, channel_id_return

bot = telebot.TeleBot(token())
n = 0
impl = 0
# Создание папки с фото
if not os.path.isdir("Photoo"):
    os.mkdir("Photoo")

os.chdir("Photoo")
ph = os.getcwd()
print(ph)


# установка эксель файла если его нет в папке с ботом
def setup_xlsx():
    try:
        open('schedule.xlsx')
    except Exception:
        wb = openpyxl.Workbook()
        wb.save('schedule.xlsx')
setup_xlsx()


if n == 0:
    n = 1
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    while ws[f'A{n}'].value is not None:
        n += 1
        if n != 0:
            n -= 1

print(n)


# команда старт
@bot.message_handler(commands=['start'])
def starts(message):
    message_handler.message_1(message, bot)


@bot.message_handler(commands=['info_mes'])
def inf_mess(message):
    bot.send_message(message.chat.id, "пришли сообщение")
    bot.register_next_step_handler(message, in_fo)
def in_fo(message):
    bot.send_message(message.chat.id, f"{message}")


# Получение id чата
@bot.message_handler(commands=['get_id'])
def starts(message):
    bot.send_message(message.chat.id, "пришлите текстовое сообщение из группы, для получения id")
    bot.register_next_step_handler(message, id_chat_reg)


def id_chat_reg(message):
    try:
        bot.send_message(message.chat.id, f"ID вашего чата:<pre>{message.forward_from_chat.id}</pre>",
                         parse_mode="HTML")
    except AttributeError:
        bot.send_message(message.chat.id, "Что-то пошло не так, попробуйте прислать сообщение из группы")
        bot.register_next_step_handler(message, id_chat_reg)


@bot.message_handler(content_types=['text'], func=lambda message: message.text == "Вернуться в главное меню")
def menu(message):
    global n
    message_handler.message_2(message, bot)


@bot.message_handler(content_types=['text'], func=lambda message: message.text == "Создать пост")
def hello_answer(message):
    global impl
    impl = 0
    bot.send_message(message.chat.id, text="Пришли фото теперь в любом формате))")


@bot.message_handler(content_types=['photo', 'document'])
def get_user_text(message):
    s = None
    df = None
    global n
    global ph
    if message.content_type == 'document':
        file_info = bot.get_file(message.document.file_id)
        df = bot.download_file(file_info.file_path)
        file_format_in = message.document.file_name.rfind('.')
        file_format = message.document.file_name[file_format_in:]
        print(df)
        s = f"{ph}/" + message.document.file_id + file_format
    elif message.content_type == 'photo':
        file_info = bot.get_file(message.photo[-1].file_id)
        df = bot.download_file(file_info.file_path)
        file_format = '.' + "PNG"
        file_unique_id = message.photo[-1].file_unique_id
        file_name = f"{file_unique_id}{file_format}"
        s = f"{ph}/{file_name}"

        # file_info = bot.get_file(message.photo.file_id)
        # df = bot.download_file(file_info.file_path)
        # file_format_in = message.photo.file_name.rfind('.')
        # file_format = message.photo.file_name[file_format_in:]
        # print(df)
        # s = f"{ph}/" + message.file_id + file_format

    src = s
    downloaded_file = df
    with open(src, 'wb') as new_file:
        new_file.write(downloaded_file)
        n = wright_name(src, n)

    bot.send_message(message.chat.id, "ВВедите текст к фото")
    bot.register_next_step_handler(message, get_message_text_post, n)
    # get_message_text_post(message, n)


def get_message_text_post(message, n):
    global impl
    try:
        wright_text(message.text, n, 10)
        message_handler.message_3(message, bot, n, impl)
    except AttributeError:
        get_message_text_post(message, n)


@bot.message_handler(content_types=['text'], func=lambda message: message.text == "❓ Количество постов")
def hello_answer(message):
    c = 1
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    while ws[f'A{c}'].value is not None:
        c += 1
    bot.send_message(message.chat.id, text=f"всего постов: {c-1}")



@bot.message_handler(commands=['photo'])
def get_user_text(message):
    global n
    global impl
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    while ws[f'A{n}'].value is not None:
        n += 1
    src = "photo"
    wright_name(src, n)
    message_handler.message_3(message, bot, n, impl)




@bot.callback_query_handler(func=lambda call: call.data.startswith("but1"))
def handle(call):
    n = call.data.replace('but1', '')
    bot.send_message(call.message.chat.id, 'введите время (12:00)'.format(str(call.data)))
    bot.answer_callback_query(call.id)
    bot.register_next_step_handler(call.message, setup, n)


def setup(message, n):
    global impl
    try:
        user_time = datetime.strptime(message.text, '%H:%M').time()
    except ValueError:
        bot.send_message(message.chat.id, "Введено некорректное время!")
        bot.register_next_step_handler(message, setup, n)
    else:
        bot.send_message(message.chat.id, f"установленное время {user_time.strftime('%H:%M')}")
        wright_time(user_time, n)
        impl = 1
        message_handler.message_3(message, bot, n, impl)


@bot.callback_query_handler(func=lambda call: call.data.startswith("but2"))
def handle(call):
    n = call.data.replace('but2', '')
    message_handler.message_4(call.message, bot, n)


# удаление поста

# Почему нельзя удалить пост если отправить его 2-раз(те нельзя удолять один пост 2-а раза
#  а точнее работает с задержкой
@bot.callback_query_handler(func=lambda call: call.data.startswith("but3"))
def dell(call):
    # n = call.data.replace('but3', '')
    global n
    print(n)
    message_id = ret_urn_day(n, 8)
    chat_id = ret_urn_day(n, 9)
    if chat_id is None:
        chat_id = channel_id_return(n)
    bot.delete_message(chat_id=chat_id, message_id=message_id)
    bot.send_message(call.message.chat.id, 'последний пост удален')


@bot.callback_query_handler(func=lambda call: call.data.startswith("buton0"))
def time_date(call):
    n = call.data.replace('buton0', '')
    bot.send_message(call.message.chat.id, 'введите дату (Д-М)'.format(str(call.data)))
    bot.answer_callback_query(call.id)
    bot.register_next_step_handler(call.message, setup0, n)


def setup0(message, n):
    try:
        user_time = datetime.strptime(message.text, '%d-%m')
    except ValueError:
        bot.send_message(message.chat.id, "Введена некорректная дата!")
        bot.register_next_step_handler(message, setup0, n)
    else:
        current_year = datetime.now().year
        full_date_user = datetime(current_year, user_time.month, user_time.day).strftime("%d-%m-%Y")
        bot.send_message(message.chat.id, f"установленная дата {full_date_user}")
        full_date = datetime.strptime(full_date_user, "%d-%m-%Y").date()
        wright_date(full_date.strftime("%d-%m-%Y"), n)
        date = ret_urn_day(n, 3)
        mons = ret_urn_day(n, 4)
        weeks = ret_urn_day(n, 5)
        days = ret_urn_day(n, 6)
        if date is not None:
            if mons is not None:
                full_date = full_date + relativedelta(months=mons)
            if weeks is not None:
                full_date = full_date + timedelta(weeks=weeks)
            if days is not None:
                full_date = full_date + timedelta(days=days)
            wright_last_days(full_date.strftime("%d-%m-%Y"), n)
        message_handler.message_4(message, bot, n)


@bot.callback_query_handler(func=lambda call: call.data.startswith("buton1"))
def time_date_month(call):
    n = call.data.replace('buton1', '')
    bot.send_message(call.message.chat.id, 'введите количество месяцов'.format(str(call.data)))
    bot.answer_callback_query(call.id)

    bot.register_next_step_handler(call.message, setup1, n)


def setup1(message, n):
    if message.text.isnumeric():
        month = int(message.text)
        bot.send_message(message.chat.id, f"установленное количество месяцов {month}")
        wright_month(month, n)
        date = ret_urn_day(n, 3)
        mons = ret_urn_day(n, 4)
        weeks = ret_urn_day(n, 5)
        days = ret_urn_day(n, 6)
        if date is None:
            date_now = datetime.now()
            date = date_now.strftime('%d-%m-%Y')
            wright_date(date, n)
        full_date = datetime.strptime(date, "%d-%m-%Y").date()
        if mons is not None:
            full_date = full_date + relativedelta(months=mons)
        if weeks is not None:
            full_date = full_date + timedelta(weeks=weeks)
        if days is not None:
            full_date = full_date + timedelta(days=days)
        wright_last_days(full_date.strftime("%d-%m-%Y"), n)

    else:
        bot.send_message(message.chat.id, "Введите число!")
        bot.register_next_step_handler(message, setup1, n)


@bot.callback_query_handler(func=lambda call: call.data.startswith("buton2"))
def time_date_week(call):
    n = call.data.replace('buton2', '')
    bot.send_message(call.message.chat.id, 'введите количество недель'.format(str(call.data)))
    bot.answer_callback_query(call.id)
    bot.register_next_step_handler(call.message, setup2, n)


def setup2(message, n):
    if message.text.isnumeric():
        week = int(message.text)
        bot.send_message(message.chat.id, f"установленное количество недель {week}")
        wright_week(week, n)

        date = ret_urn_day(n, 3)
        mons = ret_urn_day(n, 4)
        weeks = ret_urn_day(n, 5)
        days = ret_urn_day(n, 6)
        if date is None:
            date_now = datetime.now()
            date = date_now.strftime('%d-%m-%Y')
            wright_date(date, n)
        full_date = datetime.strptime(date, "%d-%m-%Y").date()
        if mons is not None:
            full_date = full_date + relativedelta(months=mons)
        if weeks is not None:
            full_date = full_date + timedelta(weeks=weeks)
        if days is not None:
            full_date = full_date + timedelta(days=days)
        wright_last_days(full_date.strftime("%d-%m-%Y"), n)
        # date = ret_urn_day(n, 3)
        # if date is None:
        #     date_now = datetime.now()
        #     full_date = date_now.strftime('%d-%m-%Y')
        #     wright_date(full_date, n)
        #     date = full_date
        # date_obj = datetime.strptime(date, '%d-%m-%Y').date()
        # result = date_obj + timedelta(weeks=week)
        # wright_last_days(result.strftime("%d-%m-%Y"), n)
    else:
        bot.send_message(message.chat.id, "Введите число!")
        bot.register_next_step_handler(message, setup2, n)


@bot.callback_query_handler(func=lambda call: call.data.startswith("buton3"))
def time_date_days(call):
    n = call.data.replace('buton3', '')
    bot.send_message(call.message.chat.id, 'введите количество дней'.format(str(call.data)))
    bot.answer_callback_query(call.id)
    bot.register_next_step_handler(call.message, setup3, n)


def setup3(message, n):
    if message.text.isnumeric():
        days = int(message.text)
        bot.send_message(message.chat.id, f"установленное количество дней {days}")
        wright_days(days, n)

        date = ret_urn_day(n, 3)
        mons = ret_urn_day(n, 4)
        weeks = ret_urn_day(n, 5)
        days = ret_urn_day(n, 6)
        if date is None:
            date_now = datetime.now()
            date = date_now.strftime('%d-%m-%Y')
            wright_date(date, n)
        full_date = datetime.strptime(date, "%d-%m-%Y").date()
        if mons is not None:
            full_date = full_date + relativedelta(months=mons)
        if weeks is not None:
            full_date = full_date + timedelta(weeks=weeks)
        if days is not None:
            full_date = full_date + timedelta(days=days)
        wright_last_days(full_date.strftime("%d-%m-%Y"), n)
        # date = ret_urn_day(n, 3)
        # if date is None:
        #     date_now = datetime.now()
        #     full_date = date_now.strftime('%d-%m-%Y')
        #     wright_date(full_date, n)
        #     date = full_date
        # date_obj = datetime.strptime(date, '%d-%m-%Y').date()
        # result = date_obj + timedelta(days=days)
        # wright_last_days(result.strftime("%d-%m-%Y"), n)
    else:
        bot.send_message(message.chat.id, "Введите число!")
        bot.register_next_step_handler(message, setup3, n)


# Вернуться в главное меню
@bot.callback_query_handler(func=lambda call: call.data.startswith("but4"))
def beck(call):
    # n = call.data.replace('but4', '')
    message_handler.message_2(call.message, bot)


# Вернуться назад
@bot.callback_query_handler(func=lambda call: call.data.startswith("buton4"))
def handle(call):
    global impl
    n = call.data.replace('buton4', '')
    message_handler.message_3(call.message, bot, n, impl)


# Проверить настройки поста
@bot.callback_query_handler(func=lambda call: call.data.startswith("but5"))
def handle(call):
    bot.send_message(call.message.chat.id, "Какой пост проверить")
    bot.register_next_step_handler(call.message, settings)
    # settings(call.message)


def settings(message):
    cnn = message.text
    times = ret_urn_day(cnn, 2)
    dwy = ret_urn_day(cnn, 3)
    dwys = ret_urn_day(cnn, 7)
    photo_import = ret_urn_day(cnn, 1)
    text_ret = ret_urn_day(cnn, 10)
    # замена обратных слешей
    photo_import = photo_import.replace("\\", "/")
    with open(photo_import, 'rb') as photo:
        bot.send_photo(message.chat.id, photo, caption=
        f"{text_ret}")
        bot.send_message(message.chat.id,
        f"время отправки поста {times}\n"
        f"Дата отправки первого поста {dwy}\n"
        f"Дата отправки второго поста {dwys}\n" )


# id чата
@bot.callback_query_handler(func=lambda call: call.data.startswith("but6"))
def handle(call):
    n = call.data.replace('but6', '')
    bot.send_message(call.message.chat.id, f'введите id чата, если не знаете id чата воспользуйтесь командой /get_id'
                     .format(str(call.data)))
    bot.answer_callback_query(call.id)
    bot.register_next_step_handler(call.message, id_chat, n)


def id_chat(message, n):
    if re.match(r'^-?\d+$', message.text):
        bot.send_message(message.chat.id, f"ID чата:{message.text}")
        wright_chat_id(message.text, n)
    else:
        bot.send_message(message.chat.id, "Введите число.")
        bot.register_next_step_handler(message, id_chat, n)
    print(n)


@bot.message_handler(content_types=['text'], func=lambda message: message.text == "👋 выложить пост")
def send_mes(message):
    global n
    channel_id = ret_urn_day(n, 9)
    print(n)
    if channel_id is None:
        channel_id = channel_id_return(n)
        # nn = n
        # a = openpyxl.load_workbook('schedule.xlsx')
        # ws = a.active
        # while ws[f'I{nn}'].value is None:
        #     nn -= 1
        #     print(nn)
        # channel_id = ws[f'I{nn}'].value
        # print(channel_id)
    print(f"channel_id{channel_id}")
    return_text = ret_urn_day(n, 10)
    print(return_text)
    try:
        print(n)
    except NameError:
        print("Сначала загрузите фото")
        bot.send_message(message.chat.id, "Сначала загрузите фото")
    else:
        try:
            print(channel_id)
        except NameError:
            bot.send_message(message.chat.id, "Сначала введите id чата")
        else:
            bot.send_message(message.chat.id, "Пост отправлен")
            send_message(return_text, channel_id)


def send_message(message_text, channel_id):
    global n
    print(n)
    send_time = "14:04"
    send_date = "2023-04-03"
    photo_import = ret_urn_day(n, 1)
    # замена обратных слешей
    photo_import = photo_import.replace("\\", "/")
    with open(photo_import, 'rb') as photo:
        sent_message = bot.send_photo(channel_id, photo, caption=
        f"{message_text}")

        message_id = sent_message.message_id
        wright_delete(message_id, n)
        print(f"Фото id: {sent_message.message_id}")

    #     def delete_message():
    #         bot.delete_message(chat_id=channel_id, message_id=message_id)
    # schedule.every(3).minutes.do(delete_message)
    # schedule.every().day.at(send_time).do(send_message, channel_id, message_text)

    # while True:
    #     schedule.run_pending()
    #     time.sleep(1)




print("bot started")
bot.infinity_polling()

# while True:
#     schedule.run_pending()
#     time.sleep(1)

# Проверка типа файла
# Инфа для фото
# @bot.message_handler(content_types=['photo'])
# def handle_photo(message):
#     print(10)
#     # получаем информацию о фото
#     file_info = bot.get_file(message.photo[-1].file_id)
#     file_url = f"https://api.telegram.org/file/bot{token()}/" \
#                f"{file_info.file_path}"
#     response = requests.get(file_url)
#     img = Image.open(BytesIO(response.content))
#
#     # выводим информацию о фото
#     print(f"Image format: {img.format}")
#     print(f"Image size: {img.size}")
