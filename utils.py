import openpyxl


def wright_name(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    while ws[f'A{n}'].value is not None:
        n += 1
    ws[f'A{n}'] = src
    print(ws[f'A{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_time(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'B{n}'] = src
    print(ws[f'B{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_date(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'C{n}'] = src
    print(ws[f'C{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_month(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'D{n}'] = src
    print(ws[f'D{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_week(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'E{n}'] = src
    print(ws[f'E{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_days(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'F{n}'] = src
    print(ws[f'F{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_last_days(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'G{n}'] = src
    print(ws[f'G{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_delete(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'H{n}'] = src
    print(ws[f'H{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_chat_id(src, n):
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'I{n}'] = src
    print(ws[f'I{n}'].value)
    a.save('schedule.xlsx')
    return n


def wright_text(src, n, b):
    mass = {"1": "A", "2": "B", "3": "C", "4": "D", "5": "E", "6": "F", "7": "G", "8": "H", "9": "I", "10": "J"}
    cell = mass[f"{b}"]
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    ws[f'{cell}{n}'] = src
    print(ws[f'{cell}{n}'].value)
    a.save('schedule.xlsx')
    return n


def ret_urn_day(n, b):
    mass = {"1": "A", "2": "B", "3": "C", "4": "D", "5": "E", "6": "F", "7": "G", "8": "H", "9": "I", "10": "J" }
    cell = mass[f"{b}"]
    a = openpyxl.load_workbook('schedule.xlsx')
    ws = a.active
    src = ws[f'{cell}{n}'].value
    return src

